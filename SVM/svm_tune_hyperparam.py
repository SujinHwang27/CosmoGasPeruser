import numpy as np
import pandas as pd

from sklearn.svm import SVC
from sklearn.model_selection import KFold, train_test_split, GridSearchCV

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

CLASS_SIZE = 16000

# TODO : show or return the best result (user can choose overfit amount?)

"""
Train SVM with k-fold cross validation and hyperparameter tuning, where k=5
"""

# TODO : delete color column, add result to existing result file
def _dump_grid_search_results_(gs_results, excel_file):    
    gs_df = {}
    for redshift, grid_search in gs_results.items():
        df = pd.DataFrame.from_dict(grid_search.cv_results_)
        
        df['overfit'] = df['mean_train_score'] - df['mean_test_score']
    
        condition_1 = df['mean_train_score'] == 1
        condition_2 = abs(df['mean_test_score'] - df['mean_train_score']) == \
                      abs(df['mean_test_score'] - df['mean_train_score']).min()
    
        df['highlight'] = ''  
        df.loc[condition_1, 'highlight'] = 'red'
        df.loc[condition_2, 'highlight'] = 'green'
    
        gs_df[redshift] = df
    
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        for redshift, df in gs_df.items():
            df.to_excel(writer, sheet_name=f'redshift_{redshift}', index=False)
    
    wb = load_workbook(excel_file)
    
    for redshift, df in gs_df.items():
        sheet = wb[f'redshift_{redshift}']
        
        for row_idx, highlight in enumerate(df['highlight'], start=2):  
            fill = None
            if highlight == 'red':
                fill = PatternFill(start_color="E6B8B7", end_color="E6B8B7", fill_type="solid")
            elif highlight == 'green':
                fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
            
            if fill:
                for col_idx in range(1, len(df.columns) + 1):  
                    sheet.cell(row=row_idx, column=col_idx).fill = fill
    
    wb.save(excel_file)



def _data_sampler_(data, size_per_class):
    X, y = [], []
    
    for physics, spectra in data.items():
        # TODO : might need consistent indices across physicses
        indices = np.random.choice(data.shape[0], size=size_per_class, replace=(CLASS_SIZE<size_per_class))  # replace should be False as long as size_per_class is less than 16000
        X.append(spectra[indices])
        y.extend([physics] * size_per_class)
    
    X = np.vstack(X)
    y = np.array(y)
    print(f'X dimension: {X.shape}, y dimension: {y.shape}')
    
    return X, y


# TODO : change file name(data size), consistency of n_comp and variance
def svm_tune_hyperparam(data, size_per_class, param_grid, n_comp, k=5):

    X, y = _data_sampler_(data, size_per_class)

    print("Performing grid search...")
    grid_search = GridSearchCV(SVC(), param_grid, cv=KFold(n_splits=k, shuffle=True), scoring='accuracy', verbose=1, return_train_score=True)
    grid_search.fit(X, y)
    # best_params = grid_search.best_params_
    # print(f"Best parameters: {best_params}")
    # print(f"Best cross-validation accuracy: {grid_search.best_score_:.4f}")
        
    return grid_search

_dump_grid_search_results_(grid_search, f"gs_results_{n_comp}_{size_per_class}.xlsx")    


def performance_vs_overfit():
    excel_file = "consolidated_gs_results.xlsx"
    sheets = pd.ExcelFile(excel_file).sheet_names  
    
    for sheet in sheets:
        df = pd.read_excel(excel_file, sheet_name=sheet)
        
        if 'overfit' not in df.columns or 'mean_test_score' not in df.columns:
            print(f"Skipping sheet {sheet}: required columns not found.")
            continue
        
        plt.figure(figsize=(8, 5))
        plt.scatter(df['mean_test_score'], df['overfit'], alpha=0.7, c='blue', label='Overfit vs Test Score')
        plt.axhline(0, color='red', linestyle='--', linewidth=1, label='No Overfit Line')
        
        plt.title(f"Overfit vs Mean Test Score for {sheet}")
        plt.xlabel("Mean Test Score")
        plt.ylabel("Overfit (Train - Test Score)")
        plt.legend()
        plt.grid(alpha=0.5)

def factors_heatmap():
    file_path = 'consolidated_gs_results.xlsx'
    sheet_names = ['redshift_0.1', 'redshift_0.3', 'redshift_2.2', 'redshift_2.4']
    
    columns_to_include = ['param_C', 'param_gamma', 'param_kernel', 'mean_test_score', 
                          'mean_train_score', 'overfit', 'n_comp', 'data_size']
    
    for sheet in sheet_names:
        data = pd.read_excel(file_path, sheet_name=sheet)
        selected_data = data[columns_to_include]
        selected_data = pd.get_dummies(selected_data, drop_first=False)  # One-hot encoding
        correlation_matrix = selected_data.corr()
        
        plt.figure(figsize=(10, 8))
        sns.heatmap(correlation_matrix, annot=True, cmap='coolwarm', fmt=".2f", cbar=True)
        plt.title(f"Factors correlation Heatmap for {sheet}")
        plt.show()

factors_heatmap()
